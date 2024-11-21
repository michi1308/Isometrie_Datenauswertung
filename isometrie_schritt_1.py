import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import pandas as pd
from openpyxl import load_workbook

# Liste der benötigten Arbeitsblätter
required_sheets = [
    "Isometr_Kon_Exz_60_5_Links",
    "Isometr_Kon_Exz_60_5_Rechts",
    "Isometr_Exz_Kon_30_5_Links",
    "Isometr_Exz_Kon_30_5_Rechts"
]

# Hilfsfunktion zur Ausgabe in das Text-Widget
def output_to_widget(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)
    text_widget.update_idletasks()  # Aktualisiert das Textfeld sofort

def excel_dateien_verarbeiten(file_path, file_name, text_widget):
    """Verarbeitet eine einzelne Excel-Datei und gibt ein Dictionary mit den gewünschten Daten zurück."""
    data = {
        'Dateiname': file_name,
        'Name': "n.a.",
        'ID': "n.a.",
        'Max Extension links': 'nachbearbeiten',
        'Max Extension rechts': 'nachbearbeiten',
        'Max Flexion links': 'nachbearbeiten',
        'Max Flexion rechts': 'nachbearbeiten',
        'Seitenunterschied Extension absolut': 'nachbearbeiten',
        'Seitenunterschied Extension relativ (%)': 'nachbearbeiten',
        'Seitenunterschied Flexion absolut': 'nachbearbeiten',
        'Seitenunterschied Flexion relativ (%)': 'nachbearbeiten',
        'Verhältnis Flexion/Extension links': 'nachbearbeiten',
        'Verhältnis Flexion/Extension rechts': 'nachbearbeiten',
        'Unterschied Extension/Flexion links': 'nachbearbeiten',
        'Unterschied Extension/Flexion rechts': 'nachbearbeiten',
        'Winkel maximales Drehmoment links Extension': 'nachbearbeiten',
        'Winkel maximales Drehmoment rechts Extension': 'nachbearbeiten',
        'Winkel maximales Drehmoment links Flexion': 'nachbearbeiten',
        'Winkel maximales Drehmoment rechts Flexion': 'nachbearbeiten'
    }

    workbook = None
    try:
        # Lade die Arbeitsmappe
        workbook = load_workbook(file_path, data_only=True)

        # Werte aus dem "Wiederholungen"-Blatt extrahieren
        if "Wiederholungen" in workbook.sheetnames:
            sheet = workbook["Wiederholungen"]
            data['Name'] = sheet["A2"].value if sheet["A2"].value else "n.a."
            data['ID'] = sheet["B2"].value if sheet["B2"].value else "n.a."
        else:
            output_to_widget(text_widget, f"{file_name}: Das Arbeitsblatt 'Wiederholungen' fehlt.")

        # Überprüfe jedes erforderliche Arbeitsblatt und finde die maximalen Werte
        for sheet_name in required_sheets:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Extrahiere die Drehmomentwerte aus Spalte C
                torque_values = [cell.value for cell in sheet['C'] if isinstance(cell.value, (int, float))]

                if torque_values:
                    max_torque = max(torque_values)
                    max_index = torque_values.index(max_torque) + 1  # Excel-Indizes beginnen bei 1

                    # Hole den zugehörigen Winkelwert aus Spalte B in derselben Zeile wie der Maximalwert
                    angle_cell = sheet[f"B{max_index + 1}"].value

                    # Speichere die Ergebnisse in das Dictionary
                    if sheet_name == "Isometr_Kon_Exz_60_5_Links":
                        data['Max Extension links'] = max_torque
                        data['Winkel maximales Drehmoment links Extension'] = angle_cell
                    elif sheet_name == "Isometr_Kon_Exz_60_5_Rechts":
                        data['Max Extension rechts'] = max_torque
                        data['Winkel maximales Drehmoment rechts Extension'] = angle_cell
                    elif sheet_name == "Isometr_Exz_Kon_30_5_Links":
                        data['Max Flexion links'] = max_torque
                        data['Winkel maximales Drehmoment links Flexion'] = angle_cell
                    elif sheet_name == "Isometr_Exz_Kon_30_5_Rechts":
                        data['Max Flexion rechts'] = max_torque
                        data['Winkel maximales Drehmoment rechts Flexion'] = angle_cell
            else:
                output_to_widget(text_widget, f"{file_name}: Das Arbeitsblatt '{sheet_name}' fehlt.")

        # Berechnungen der Seitenunterschiede und Verhältnisse
        max_extension_links = data['Max Extension links']
        max_extension_rechts = data['Max Extension rechts']
        max_flexion_links = data['Max Flexion links']
        max_flexion_rechts = data['Max Flexion rechts']

        # Seitenunterschied und Verhältnisse für Extension
        if isinstance(max_extension_links, (int, float)) and isinstance(max_extension_rechts, (int, float)):
            seitenunterschied_extension_absolut = abs(max_extension_links - max_extension_rechts)
            min_extension = min(max_extension_links, max_extension_rechts)
            max_extension = max(max_extension_links, max_extension_rechts)
            seitenunterschied_extension_relativ = round((1 - (min_extension / max_extension)) * 100, 2)
        else:
            seitenunterschied_extension_absolut = "nachbearbeiten"
            seitenunterschied_extension_relativ = "nachbearbeiten"

        # Seitenunterschied und Verhältnisse für Flexion
        if isinstance(max_flexion_links, (int, float)) and isinstance(max_flexion_rechts, (int, float)):
            seitenunterschied_flexion_absolut = abs(max_flexion_links - max_flexion_rechts)
            min_flexion = min(max_flexion_links, max_flexion_rechts)
            max_flexion = max(max_flexion_links, max_flexion_rechts)
            seitenunterschied_flexion_relativ = round((1 - (min_flexion / max_flexion)) * 100, 2)
        else:
            seitenunterschied_flexion_absolut = "nachbearbeiten"
            seitenunterschied_flexion_relativ = "nachbearbeiten"

        # Verhältnis für Flexion/Extension
        verhaeltnis_flexion_extension_links = round(max_flexion_links / max_extension_links,
                                                    2) if max_extension_links else "nachbearbeiten"
        verhaeltnis_flexion_extension_rechts = round(max_flexion_rechts / max_extension_rechts,
                                                     2) if max_extension_rechts else "nachbearbeiten"

        # Unterschied für Extension/Flexion
        unterschied_extension_flexion_links = abs(
            max_extension_links - max_flexion_links) if max_extension_links and max_flexion_links else "nachbearbeiten"
        unterschied_extension_flexion_rechts = abs(
            max_extension_rechts - max_flexion_rechts) if max_extension_rechts and max_flexion_rechts else "nachbearbeiten"

        # Hinzufügen der berechneten Werte zum Daten-Dictionary
        data.update({
            'Seitenunterschied Extension absolut': seitenunterschied_extension_absolut,
            'Seitenunterschied Extension relativ (%)': seitenunterschied_extension_relativ,
            'Seitenunterschied Flexion absolut': seitenunterschied_flexion_absolut,
            'Seitenunterschied Flexion relativ (%)': seitenunterschied_flexion_relativ,
            'Verhältnis Flexion/Extension links': verhaeltnis_flexion_extension_links,
            'Verhältnis Flexion/Extension rechts': verhaeltnis_flexion_extension_rechts,
            'Unterschied Extension/Flexion links': unterschied_extension_flexion_links,
            'Unterschied Extension/Flexion rechts': unterschied_extension_flexion_rechts
        })

    except Exception as e:
       output_to_widget(text_widget, f"Fehler beim Verarbeiten der Datei {file_name}: {e}")

    finally:
        # Sicherstellen, dass die Arbeitsmappe geschlossen wird
        if workbook:
            workbook.close()

    return data


# Hauptausführungsblock
def datenverarbeitung_steuern_und_speichern(text_widget):
    # Alle Dateien im Verzeichnis verarbeiten und Ergebnisse speichern
    result_data = []
    output_to_widget(text_widget, "Datenverarbeitung gestartet...")
    for file_name in os.listdir(path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(path, file_name)
            result_data.append(excel_dateien_verarbeiten(file_path, file_name, text_widget))

    # Erstelle DataFrame und runde numerische Werte auf 2 Dezimalstellen
    result_df = pd.DataFrame(result_data)
    result_df = result_df.round(2)  # Rundet alle numerischen Werte auf 2 Nachkommastellen

    # Speichern der Ergebnisse in eine Excel-Datei
    output_file_path = os.path.join(path, "Ergebnisse_isometrisch.xlsx")
    result_df.to_excel(output_file_path, index=False)

    # Spaltenbreite auf 20 setzen
    workbook = load_workbook(output_file_path)
    try:
        sheet = workbook.active
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 20  # Breite auf 20 setzen
        workbook.save(output_file_path)
    finally:
        workbook.close()  # Sicherstellen, dass die Arbeitsmappe immer geschlossen wird

    output_to_widget(text_widget, f"Die Ergebnisse wurden erfolgreich in {output_file_path} gespeichert.")
    messagebox.showinfo("Erfolg",
                        f"Die Ergebnistabelle wurde erfolgreich erstellt und gespeichert unter: {output_file_path}")

# GUI-Setup
def ordner_auswaehlen(entry):
    folder_selected = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, folder_selected)

def verabeitung_starten(entry, text_widget):
    folder_path = entry.get()
    if not os.path.isdir(folder_path):
        messagebox.showerror("Fehler", "Bitte geben Sie einen gültigen Ordnerpfad an.")
        return
    text_widget.delete(1.0, tk.END)  # Löscht die Textausgabe
    datenverarbeitung_steuern_und_speichern(text_widget)

def main():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("Isometrie Datenauswertung")

    # Eingabefeld für Ordnerpfad
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    entry_label = tk.Label(frame, text="Bitte den Ordnerpfad angeben:")
    entry_label.grid(row=0, column=0, sticky="w")

    # Einstellen der Breite des Eingabefeldes
    entry = tk.Entry(frame, width=70)
    entry.grid(row=0, column=1)

    browse_button = tk.Button(frame, text="Durchsuchen", command=lambda: ordner_auswaehlen(entry))
    browse_button.grid(row=0, column=2, padx=5)

    # Start-Button
    start_button = tk.Button(frame, text="Starten", command=lambda: verabeitung_starten(entry, text_output))
    start_button.grid(row=1, column=1, pady=10)

    # Text-Widget für die Ausgaben
    text_output = ScrolledText(root, height=20, width=100)
    text_output.pack(padx=10, pady=10)

    root.mainloop()

# Hauptprogramm starten
if __name__ == "__main__":
    main()



